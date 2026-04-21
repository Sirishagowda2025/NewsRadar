"""
NewsRadar — AI-Powered Industry Intelligence Monitor  v2.0
===========================================================
Fetches news from Google News RSS → filters → AI summary → generates HTML.

HOW TO RUN:
    # 1. pip install ollama feedparser openpyxl python-dotenv requests
    # 2. Make sure Ollama desktop app is open (or: ollama serve)
    # 3. Copy .env.example → .env and fill in your values
    # 4. Run:
    python main.py                        # default industry (e-commerce)
    python main.py --industry fintech     # fintech profile
    python main.py --industry healthcare  # healthcare profile
    python main.py --send-email           # send last report by email

NEW IN v2.0:
    FEATURE 1 — Multi-industry profiles via --industry flag.
                Built-in: ecommerce (default), fintech, healthcare, tech.
                Add your own in INDUSTRY_PROFILES dict in config.py.
    FEATURE 2 — Filter bar in the HTML report: filter by AI score (5+/7+/9+)
                and by date (today / this week / all). Instant, no reload.
    FEATURE 3 — Dark mode toggle button in the topbar. Preference is kept
                for the session.
    FEATURE 4 — Slack webhook: post digest summaries to a Slack channel.
                Set SLACK_WEBHOOK_URL in .env to enable.
    FEATURE 5 — Live search bar in the report. Type to filter cards across
                all sections in real time.

OUTPUT:
    newsradar_report.html   <- open in browser, review, drag/remove
    newsradar_report.xlsx          <- full data export

PIPELINE PER ARTICLE:
    Gate 0  ->  skip irrelevant source domains  (IRRELEVANT_SOURCES list)
    Gate 1  ->  skip articles with blocked keywords  (EXCLUDE_KEYWORDS)
    Gate 2  ->  must mention at least 1 platform/keyword  (RELEVANCE_KEYWORDS)
    Dedup   ->  fuzzy headline (78%) + exact URL + snippet fingerprint
    Gate 3  ->  qwen3:4b relevance score via ollama Python library
    AI      ->  qwen3:4b generates 4-point summary per article

v11 FIXES vs v10:
    FIX 1 — HTML always served via local HTTP server (never file:// URI) so
             Chrome Incognito can fetch the page and all fetch() calls work.
    FIX 2 — Remove button: card removal no longer relies on transitionend
             (which silently misfires); uses a guaranteed setTimeout fallback.
    FIX 3 — Send Email: _make_email_html() now uses depth-aware regex that
             correctly strips nested card-actions divs; topbar in email
             shows ONLY the brand title (no counts, no badges, no button).
    FIX 4 — CPU: num_predict capped at 60 for scoring (only needs an integer),
             400 kept for summaries; num_thread set to half of os.cpu_count()
             so Ollama never saturates all cores.
    FIX 5 — Editable header: title stored in a module-level JS variable
             (not localStorage, which is blocked in Incognito); value is
             baked into the email HTML at send-time via /send-email endpoint
             reading the live DOM title from the POST body.
    FIX 6 — Email topbar: only brand name + date shown; all interactive
             elements (count badge, feedback badge, send button) removed.
    FIX 7 — Cross-category drag-and-drop: complete JS rewrite.
             A single document-level dragover handler covers every pixel of
             every section. Cards can be dropped onto other cards (above/below),
             onto the section dropzone strip (prepend), or onto empty section
             bodies. The card's accent colour updates to the destination section
             colour on drop. Works in Chrome Incognito and all modern browsers.
"""

import os
import sys
import feedparser
import re
import json
import time
import threading
from datetime import datetime, timedelta
from urllib.parse import quote_plus
from collections import defaultdict
from difflib import SequenceMatcher
import smtplib
from http.server import HTTPServer, BaseHTTPRequestHandler
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Optional: load .env file automatically if python-dotenv is installed
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

# Optional: requests for Slack webhook
try:
    import requests as _requests
    _REQUESTS_OK = True
except ImportError:
    _REQUESTS_OK = False

from config import (
    EMAIL_CONFIG, SYSTEM_CONFIG, OLLAMA_CONFIG, CATEGORIES,
    IRRELEVANT_SOURCES, is_irrelevant_source,
    RELEVANCE_KEYWORDS, EXCLUDE_KEYWORDS,
    TRADE_IMPACT_KEYWORDS, has_trade_impact,
)

# ─────────────────────────────────────────────────────────────────────────────
# FEATURE 1 — MULTI-INDUSTRY PROFILES
# Each profile overrides CATEGORIES, RELEVANCE_KEYWORDS, EXCLUDE_KEYWORDS.
# The --industry flag selects which profile to use at runtime.
# ─────────────────────────────────────────────────────────────────────────────

INDUSTRY_PROFILES = {
    "ecommerce": {
        "label":    "E-Commerce India",
        "categories": CATEGORIES,          # uses config.py as-is
        "relevance": RELEVANCE_KEYWORDS,
        "exclude":   EXCLUDE_KEYWORDS,
    },
    "fintech": {
        "label": "Fintech & Payments",
        "categories": {
            "Payments & UPI": {
                "html_section": "s-payments", "color": "#2563eb",
                "description": "UPI, RBI, payment gateway policy changes",
                "search_queries": [
                    "UPI transaction limit change 2026", "RBI payment regulation 2026",
                    "payment gateway fee change India 2026", "NPCI UPI policy update 2026",
                    "fintech RBI compliance mandate 2026", "digital payment fraud regulation India 2026",
                    "Razorpay Paytm PhonePe policy update 2026", "BBPS biller compliance 2026",
                ],
            },
            "Lending & Credit": {
                "html_section": "s-lending", "color": "#7c3aed",
                "description": "NBFC rules, digital lending, credit policy changes",
                "search_queries": [
                    "RBI digital lending guidelines 2026", "NBFC compliance mandate 2026",
                    "credit card fee change India 2026", "buy now pay later regulation India 2026",
                    "microfinance regulation change 2026", "personal loan rate change India 2026",
                ],
            },
            "Crypto & Regulations": {
                "html_section": "s-crypto", "color": "#d97706",
                "description": "Crypto tax, SEBI, digital assets regulation",
                "search_queries": [
                    "crypto regulation India 2026", "SEBI digital asset rule 2026",
                    "crypto tax TDS India 2026", "VDA virtual digital asset rule India 2026",
                    "India CBDC digital rupee update 2026",
                ],
            },
        },
        "relevance": [
            "upi", "rbi", "payment gateway", "fintech", "nbfc", "digital lending",
            "credit card", "neft", "rtgs", "imps", "npci", "bbps", "sebi",
            "crypto", "regulation", "compliance", "fee change", "policy update",
        ],
        "exclude": [
            "ipo", "share price", "quarterly earnings", "cricket", "bollywood",
            "celebrity", "job opening", "how to invest", "mutual fund tips",
        ],
    },
    "healthcare": {
        "label": "Healthcare & Pharma",
        "categories": {
            "Drug Regulation": {
                "html_section": "s-drugs", "color": "#dc2626",
                "description": "CDSCO approvals, drug pricing, pharma compliance",
                "search_queries": [
                    "CDSCO drug approval India 2026", "drug price control order India 2026",
                    "pharma compliance mandate India 2026", "NPPA price revision 2026",
                    "medical device regulation India 2026", "generic drug policy India 2026",
                ],
            },
            "Health Tech": {
                "html_section": "s-healthtech", "color": "#059669",
                "description": "Telemedicine, health data, digital health policy",
                "search_queries": [
                    "telemedicine regulation India 2026", "ABDM health data policy 2026",
                    "digital health compliance India 2026", "Ayushman Bharat policy change 2026",
                    "health insurance regulation IRDAI 2026",
                ],
            },
        },
        "relevance": [
            "cdsco", "drug", "pharma", "nppa", "medical device", "telemedicine",
            "health data", "abdm", "ayushman", "irdai", "compliance", "approval",
            "regulation", "price control", "mandate",
        ],
        "exclude": [
            "ipo", "share price", "cricket", "bollywood", "celebrity",
            "job opening", "fitness tips", "diet", "recipe",
        ],
    },
    "tech": {
        "label": "Technology & AI",
        "categories": {
            "AI & LLMs": {
                "html_section": "s-ai", "color": "#7c3aed",
                "description": "AI regulation, model releases, LLM policy changes",
                "search_queries": [
                    "AI regulation policy India 2026", "LLM model release update 2026",
                    "OpenAI Anthropic Google AI update 2026", "EU AI Act compliance 2026",
                    "India AI policy mandate 2026", "AI copyright regulation 2026",
                ],
            },
            "Cloud & SaaS": {
                "html_section": "s-cloud", "color": "#0891b2",
                "description": "Cloud pricing, data localisation, SaaS compliance",
                "search_queries": [
                    "AWS Azure GCP pricing change 2026", "data localisation rule India 2026",
                    "cloud compliance mandate India 2026", "SaaS regulation India 2026",
                    "CERT-In compliance deadline 2026",
                ],
            },
            "Cybersecurity": {
                "html_section": "s-security", "color": "#dc2626",
                "description": "Data breach, CERT-In, cybersecurity policy",
                "search_queries": [
                    "CERT-In cybersecurity directive 2026", "data breach regulation India 2026",
                    "DPDP Act compliance deadline 2026", "cybersecurity mandate India 2026",
                ],
            },
        },
        "relevance": [
            "ai regulation", "llm", "openai", "anthropic", "google ai",
            "cloud pricing", "aws", "azure", "gcp", "data localisation",
            "cert-in", "dpdp", "cybersecurity", "compliance", "policy update",
            "mandate", "regulation", "fee change",
        ],
        "exclude": [
            "ipo", "share price", "cricket", "bollywood", "celebrity",
            "job opening", "how to code", "tutorial", "beginner guide",
        ],
    },
}


def load_industry_profile(industry: str) -> dict:
    """Load the industry profile for the given --industry flag value."""
    key = industry.lower().strip()
    if key not in INDUSTRY_PROFILES:
        print(f"[INDUSTRY] Unknown profile '{industry}'. Available: {', '.join(INDUSTRY_PROFILES)}")
        print(f"[INDUSTRY] Falling back to 'ecommerce'")
        key = "ecommerce"
    profile = INDUSTRY_PROFILES[key]
    print(f"[INDUSTRY] Loaded profile: {profile['label']}")
    return profile

# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL STATE
# ─────────────────────────────────────────────────────────────────────────────

SEEN_HEADLINES = set()
SEEN_SNIPPETS  = set()
SEEN_URLS      = set()


# ─────────────────────────────────────────────────────────────────────────────
# AI — ollama Python library
# ─────────────────────────────────────────────────────────────────────────────

_AI_READY = None   # None = untested | True = OK | False = unavailable


def _check_ai() -> bool:
    global _AI_READY
    if _AI_READY is not None:
        return _AI_READY

    print("\n" + "-" * 60)
    print("[AI] Checking Ollama connection...")

    if not OLLAMA_CONFIG.get("enabled", True):
        _AI_READY = False
        print("[AI] SKIP — 'enabled' is False in OLLAMA_CONFIG.")
        print("-" * 60 + "\n")
        return False

    try:
        import ollama as _ollama_lib
    except ImportError:
        _AI_READY = False
        print("[AI] SKIP — ollama package not installed.")
        print("     Run:  pip install ollama")
        print("-" * 60 + "\n")
        return False

    model = OLLAMA_CONFIG.get("model", "qwen3:4b")

    try:
        models_response = _ollama_lib.list()
        available = [m.model for m in models_response.models]
        matched = [m for m in available if m.startswith(model.split(":")[0])]

        if not matched:
            _AI_READY = False
            print(f"[AI] SKIP — model '{model}' not found in Ollama.")
            print(f"     Available: {available if available else 'none'}")
            print(f"     Run:  ollama pull {model}")
            print("-" * 60 + "\n")
            return False

        _AI_READY = True
        print(f"[AI] Connected — using model: {matched[0]}")
        print("-" * 60 + "\n")
        return True

    except Exception as e:
        _AI_READY = False
        print(f"[AI] SKIP — cannot reach Ollama.")
        print(f"     Error: {type(e).__name__}: {e}")
        print("     Is Ollama running? Start with:  ollama serve")
        print("-" * 60 + "\n")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# AI HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _normalise(text, max_len=600):
    if not text:
        return ""
    for src, dst in [("&amp;","&"),("&lt;","<"),("&gt;",">"),
                     ("&quot;",'"'),("&#39;","'"),("&nbsp;"," ")]:
        text = text.replace(src, dst)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:max_len]


def _extract_json(raw):
    if not raw:
        return ""
    raw = re.sub(r"<think>.*?</think>", "", raw, flags=re.DOTALL)
    raw = re.sub(r"<think>.*",          "", raw, flags=re.DOTALL)
    raw = re.sub(r"```[a-zA-Z]*\n?",    "", raw).strip().rstrip("`").strip()
    m = re.search(r"\{[^{}]*\}", raw, re.DOTALL)
    if m:
        return m.group(0).strip()
    m = re.search(r"\[.*?\]", raw, re.DOTALL)
    if m:
        return m.group(0).strip()
    return raw.strip()


def _flatten_summary(data):
    if isinstance(data, list):
        if all(isinstance(x, str) for x in data):
            return [s for s in data if s.strip()][:5]
        out = []
        for item in data:
            if isinstance(item, dict):
                out.extend(str(v) for v in item.values() if str(v).strip())
            elif isinstance(item, str) and item.strip():
                out.append(item)
        return out[:5]
    if isinstance(data, dict):
        return [str(v) for v in data.values() if str(v).strip()][:5]
    return []


def _score_from_prose(raw):
    m = re.search(r"(?:score|rating|relevance)\D{0,10}(\d{1,2})", raw, re.IGNORECASE)
    if m:
        return max(0, min(10, int(m.group(1))))
    digits = re.findall(r"\b(10|[0-9])\b", raw)
    if digits:
        return max(0, min(10, int(digits[0])))
    return 5


def _cpu_thread_count():
    """Use at most half the logical CPUs so Ollama never saturates the machine."""
    try:
        cpus = os.cpu_count() or 4
        return max(2, cpus // 2)
    except Exception:
        return 2


def _llm_chat(user_content, prefill="{", max_tokens=400):
    """
    FIX 4 — num_predict is now a parameter.
    Scoring calls pass 60; summary calls pass 400.
    num_thread is capped to half of cpu_count.
    """
    if not _check_ai():
        return ""
    try:
        import ollama as _ol
        model   = OLLAMA_CONFIG.get("model", "qwen3:4b")
        options = {
            "temperature":    0.05,
            "num_predict":    max_tokens,   # FIX 4: caller controls token budget
            "repeat_penalty": 1.05,
            "num_thread":     _cpu_thread_count(),  # FIX 4: cap CPU cores used
            "stop":           ["\n\n", "<|", "```"],
        }
        msgs = [
            {"role": "user",      "content": user_content},
            {"role": "assistant", "content": prefill},
        ]
        try:
            resp = _ol.chat(model=model, messages=msgs, options=options, think=False)
        except TypeError:
            resp = _ol.chat(model=model, messages=msgs, options=options)
        return prefill + resp.message.content
    except Exception as e:
        print(f"    [AI CHAT ERR] {type(e).__name__}: {e}")
        return ""


def _llm_generate(prompt):
    return _llm_chat(prompt, prefill="{")


# ─────────────────────────────────────────────────────────────────────────────
# FEEDBACK LEARNING
# ─────────────────────────────────────────────────────────────────────────────

FEEDBACK_FILE = "newsradar_feedback.json"
FEEDBACK_PORT = 5765

_FEEDBACK_EXAMPLES = None
_feedback_server   = None


def _read_feedback_file():
    if not os.path.exists(FEEDBACK_FILE):
        return []
    try:
        with open(FEEDBACK_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return []


def _write_feedback_file(entries):
    tmp = FEEDBACK_FILE + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(entries, f, indent=2, ensure_ascii=False)
    os.replace(tmp, FEEDBACK_FILE)


def _record_feedback(link, headline, source, ai_score, action):
    entries = _read_feedback_file()
    entries = [e for e in entries if e.get("link") != link]
    if action is not None:
        entries.append({
            "headline": headline,
            "source":   source,
            "link":     link,
            "ai_score": ai_score,
            "action":   action,
            "ts":       datetime.now().isoformat(),
        })
    _write_feedback_file(entries)


class _FeedbackHandler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_response(204)
        self._cors()
        self.end_headers()

    def do_GET(self):
        # FIX 1 — Serve the HTML report over HTTP so Chrome Incognito works.
        # The browser always hits http://127.0.0.1:5765/ — no file:// URIs.
        if self.path in ("/", "/index.html"):
            self.send_response(200)
            self._cors()
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.end_headers()
            try:
                html_file = "newsradar_report.html"
                if not os.path.exists(html_file):
                    raise FileNotFoundError(f"{html_file} not found")
                with open(html_file, "r", encoding="utf-8") as f:
                    self.wfile.write(f.read().encode("utf-8"))
            except Exception as exc:
                self.wfile.write(f"<h1>Error</h1><pre>{exc}</pre>".encode("utf-8"))
            return

        if self.path == "/ping":
            self.send_response(200)
            self._cors()
            self.send_header("Content-Type", "text/plain")
            self.end_headers()
            self.wfile.write(b"ok")

        elif self.path == "/feedback_export":
            try:
                entries = _read_feedback_file()
                payload = json.dumps(entries, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self._cors()
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(payload)))
                self.end_headers()
                self.wfile.write(payload)
            except Exception as exc:
                print(f"  [FEEDBACK EXPORT ERR] {exc}")
                self.send_response(500)
                self._cors()
                self.end_headers()

        else:
            self.send_response(404)
            self._cors()
            self.end_headers()

    def do_POST(self):
        if self.path == "/feedback":
            try:
                length   = int(self.headers.get("Content-Length", 0))
                body     = self.rfile.read(length)
                data     = json.loads(body)
                link     = data.get("link", "").strip()
                headline = data.get("headline", "").strip()
                source   = data.get("source", "").strip()
                ai_score = data.get("ai_score", "")
                action   = data.get("action")
                if link:
                    _record_feedback(link, headline, source, ai_score, action)
                    verb = action.upper() if action else "UNMARK"
                    print(f"  [FEEDBACK] {verb:7} — {headline[:70]}")
                self.send_response(200)
                self._cors()
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(b'{"ok":true}')
            except Exception as exc:
                print(f"  [FEEDBACK ERR] {exc}")
                self.send_response(500)
                self._cors()
                self.end_headers()

        elif self.path == "/send-email":
            # FIX 3 + FIX 5 — Accept the current report title and full HTML
            # DOM snapshot from the browser so the email reflects what the user
            # actually sees after reviewing (reorder, remove, title edit).
            try:
                length     = int(self.headers.get("Content-Length", 0))
                body       = self.rfile.read(length)
                data       = json.loads(body)
                report_title = data.get("title", "Industry Intelligence Report").strip()
                raw_html     = data.get("html", "").strip()

                if not raw_html:
                    # Fallback: read file if browser sent no HTML
                    html_file = "newsradar_report.html"
                    with open(html_file, "r", encoding="utf-8") as f:
                        raw_html = f.read()

                email_html = _make_email_html(raw_html, report_title)
                ok = send_email(email_html, label="final")
                if ok:
                    _autosave_sent_feedback(email_html)
                payload = json.dumps({"ok": ok, "error": "" if ok else "SMTP failed — check terminal"}).encode()
                self.send_response(200)
                self._cors()
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(payload)
            except Exception as exc:
                print(f"  [SEND-EMAIL ERR] {exc}")
                payload = json.dumps({"ok": False, "error": str(exc)}).encode()
                self.send_response(500)
                self._cors()
                self.send_header("Content-Type", "application/json")
                self.end_headers()
                self.wfile.write(payload)

        else:
            self.send_response(404)
            self._cors()
            self.end_headers()

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin",  "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def log_message(self, fmt, *args):
        pass


def start_feedback_server():
    global _feedback_server
    try:
        server = HTTPServer(("127.0.0.1", FEEDBACK_PORT), _FeedbackHandler)
        t = threading.Thread(target=server.serve_forever, daemon=False)
        t.start()
        _feedback_server = server
        print(f"[FEEDBACK] Server running on http://127.0.0.1:{FEEDBACK_PORT}")
        print(f"[FEEDBACK] Keep/Remove clicks will write to {FEEDBACK_FILE} instantly")
        return server
    except OSError as e:
        print(f"[FEEDBACK] Could not start server on port {FEEDBACK_PORT}: {e}")
        return None


def _load_feedback_examples():
    global _FEEDBACK_EXAMPLES
    if _FEEDBACK_EXAMPLES is not None:
        return _FEEDBACK_EXAMPLES

    _FEEDBACK_EXAMPLES = {"keep": [], "remove": []}
    entries = _read_feedback_file()
    if not entries:
        print("[LEARN] No feedback file yet — AI uses default scoring rules")
        return _FEEDBACK_EXAMPLES

    seen = {}
    for e in entries:
        if e.get("link"):
            seen[e["link"]] = e
    deduped = list(seen.values())

    kept    = sorted([e for e in deduped if e.get("action") == "keep"],
                     key=lambda x: x.get("ts",""), reverse=True)
    removed = sorted([e for e in deduped if e.get("action") == "remove"],
                     key=lambda x: x.get("ts",""), reverse=True)

    _FEEDBACK_EXAMPLES["keep"]   = [e["headline"] for e in kept[:8]   if e.get("headline")]
    _FEEDBACK_EXAMPLES["remove"] = [e["headline"] for e in removed[:8] if e.get("headline")]

    k = len(_FEEDBACK_EXAMPLES["keep"])
    r = len(_FEEDBACK_EXAMPLES["remove"])
    print(f"[LEARN] Loaded feedback: {k} keep examples, {r} remove examples → injecting into AI prompt")
    return _FEEDBACK_EXAMPLES


def _build_feedback_section():
    fb = _load_feedback_examples()
    lines = []
    if fb["keep"]:
        lines.append("Examples this user has KEPT in past reviews (score similar articles HIGH):")
        for h in fb["keep"]:
            lines.append(f"  + {h[:120]}")
    if fb["remove"]:
        if lines:
            lines.append("")
        lines.append("Examples this user has REMOVED in past reviews (score similar articles LOW):")
        for h in fb["remove"]:
            lines.append(f"  - {h[:120]}")
    if not lines:
        return ""
    return "\n\nUser preference examples from past reviews:\n" + "\n".join(lines) + "\n"


def _autosave_sent_feedback(email_html: str) -> None:
    pairs = re.findall(
        r'<a[^>]+href="([^"]+)"[^>]+class="article-headline"[^>]*>([^<]+)<',
        email_html
    )
    if not pairs:
        return
    existing   = _read_feedback_file()
    seen_links = {e.get("link","") for e in existing}
    added = 0
    ts    = datetime.now().isoformat()
    for link, headline in pairs:
        link = link.strip(); headline = headline.strip()
        if link and link not in seen_links:
            existing.append({
                "headline": headline, "source": "", "link": link,
                "ai_score": "", "action": "keep", "ts": ts, "auto": True,
            })
            seen_links.add(link)
            added += 1
    if added:
        _write_feedback_file(existing)
        print(f"[LEARN] Auto-saved {added} sent articles as 'keep' → {FEEDBACK_FILE}")


# ─────────────────────────────────────────────────────────────────────────────
# AI PROMPTS
# ─────────────────────────────────────────────────────────────────────────────

_SCORE_PROMPT = (
    "You work for an e-commerce operations team "
    "operating across Amazon India, Flipkart, Meesho, Noon, Zepto, Blinkit, "
    "Swiggy Instamart, and cross-border channels.\n\n"
    "Score the relevance of this article to your business (0-10).\n\n"
    "Score 8-10 (directly actionable):\n"
    "  - Marketplace fee, commission, or referral rate change\n"
    "  - Seller policy, KYC, listing, or account health update\n"
    "  - Compliance mandate: BIS, FSSAI, EPR, GST, customs, duty\n"
    "  - Competitor update: Unicommerce, Shiprocket, Delhivery, EasyEcom, Vinculum\n"
    "  - Logistics or last-mile delivery rate/policy change\n"
    "  - API, integration, or platform technical update\n"
    "  - Payment gateway or RBI cross-border payment update\n"
    "  - Quick-commerce policy: Zepto, Blinkit, Instamart, BigBasket\n"
    "  - ONDC policy or network participant update\n"
    "  - India FTA, trade agreement, tariff change, export incentive, DGFT, CBIC\n"
    "  - Cross-border e-commerce export/import rule change\n"
    "  - Shipping rate change, freight index move, port disruption that raises\n"
    "    India export/import cost or increases delivery time\n"
    "  - Trade war, sanctions, or route disruption with direct India shipping impact\n\n"
    "Score 5-7 (useful context):\n"
    "  - General Indian e-commerce market growth or forecast\n"
    "  - Supply chain or logistics infrastructure expansion\n"
    "  - India trade policy with indirect e-commerce impact\n"
    "  - Advertising cost or sponsored listing trends on marketplaces\n"
    "  - Cross-border logistics company news relevant to India\n"
    "  - Global freight trends that may affect India in coming weeks\n\n"
    "Score 0-4 (not relevant — drop these):\n"
    "  - Stock prices, IPO, fundraising, quarterly earnings\n"
    "  - Celebrity, sports, cricket, Bollywood\n"
    "  - Packaging material forecasts (bubble wrap, roller trucks, corrugated)\n"
    "  - Aqua farming, agriculture unrelated to e-commerce\n"
    "  - General WTO talks with no specific e-commerce clause\n"
    "  - Global statistics listicles (Statista, Forbes top-N lists)\n"
    "  - How-to guides for starting an e-commerce store\n"
    "  - Africa/global market size reports with no India angle\n\n"
    "Headline : {headline}\n"
    "Source   : {source}\n"
    "Snippet  : {snippet}\n"
    "Category : {category}\n"
    "{feedback}\n"
    'Reply with ONLY this JSON and nothing else:\n{{"score": <integer 0-10>, "reason": "<one sentence>"}}'
)

_SUMMARY_PROMPT = (
    "Summarise this e-commerce news article in exactly 4 short bullet points "
    "for an e-commerce middleware operator. Each point must be a specific fact. "
    "Include numbers, dates, platform names, fees, or percentages where available.\n\n"
    "Headline : {headline}\n"
    "Source   : {source}\n"
    "Snippet  : {snippet}\n\n"
    'Reply with ONLY a JSON array of exactly 4 plain strings:\n'
    '["<fact 1>", "<fact 2>", "<fact 3>", "<fact 4>"]'
)


# ─────────────────────────────────────────────────────────────────────────────
# PUBLIC AI FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def ai_score_article(headline, source, snippet, category):
    """
    Returns (score 0-10, reason).
    FIX 4: scoring uses max_tokens=60 — a score + one sentence fits in ~40 tokens.
    """
    if not _check_ai():
        return (10, "AI unavailable")

    h = _normalise(headline, 250)
    s = _normalise(source,    80)
    p = _normalise(snippet,  450)

    feedback_section = _build_feedback_section()
    raw = _llm_chat(
        _SCORE_PROMPT.format(headline=h, source=s, snippet=p,
                             category=category, feedback=feedback_section),
        prefill="{",
        max_tokens=60,   # FIX 4: scoring needs very few tokens
    )

    for suffix in ["}", "\"}", "\", \"reason\": \"ok\"}"]:
        for candidate in [raw, raw.rstrip() + suffix]:
            try:
                obj = json.loads(candidate)
                score = max(0, min(10, int(obj["score"])))
                return (score, str(obj.get("reason", "")))
            except Exception:
                pass

    cleaned = _extract_json(raw)
    if cleaned:
        for suffix in ["}", "\"}"]:
            for candidate in [cleaned, cleaned.rstrip() + suffix]:
                try:
                    obj = json.loads(candidate)
                    score = max(0, min(10, int(obj["score"])))
                    return (score, str(obj.get("reason", "")))
                except Exception:
                    pass

    score = _score_from_prose(raw)
    print(f"    [AI PROSE FALLBACK] score={score} | {headline[:55]}...")
    return (score, "extracted from prose")


def ai_summarise_article(headline, source, snippet, category=""):
    """
    Returns list of up to 4 bullet strings.
    FIX 4: summaries use max_tokens=400 (unchanged — needs full sentences).
    """
    if not _check_ai():
        return []

    h = _normalise(headline, 250)
    s = _normalise(source,    80)
    p = _normalise(snippet,  600)

    prompt  = _SUMMARY_PROMPT.format(headline=h, source=s, snippet=p)
    raw     = _llm_chat(prompt, prefill="[", max_tokens=400)

    for suffix in ["]", "\"]"]:
        for candidate in [raw, raw.rstrip() + suffix]:
            try:
                result = _flatten_summary(json.loads(candidate))
                if result:
                    return result
            except Exception:
                pass

    cleaned = _extract_json(raw)
    if cleaned:
        for suffix in ["]", "\"]"]:
            for candidate in [cleaned, cleaned.rstrip() + suffix]:
                try:
                    result = _flatten_summary(json.loads(candidate))
                    if result:
                        return result
                except Exception:
                    pass

    print(f"    [AI SUMMARY ERR] {headline[:55]}... | raw={repr(raw[:70])}")
    return []


# ─────────────────────────────────────────────────────────────────────────────
# DEDUPLICATION
# ─────────────────────────────────────────────────────────────────────────────

def norm(text):
    t = text.lower()
    t = re.sub(r'[-|].*$', '', t)
    t = re.sub(r'\b\d+\b', '', t)
    t = re.sub(r'[^\w\s]', '', t)
    return re.sub(r'\s+', ' ', t).strip()


def _snippet_fp(snippet):
    return norm(snippet)[:120]


def is_duplicate(headline, url, snippet=""):
    nh = norm(headline)
    nu = url.strip().lower()
    sf = _snippet_fp(snippet)

    if nu and nu in SEEN_URLS:
        return True
    if sf and len(sf) > 40 and sf in SEEN_SNIPPETS:
        return True
    for sh in SEEN_HEADLINES:
        ratio = SequenceMatcher(None, nh, sh).ratio()
        if ratio > 0.78:
            return True
        if len(nh) > 20 and len(sh) > 20 and (nh in sh or sh in nh):
            return True
    return False


def mark_seen(headline, url, snippet=""):
    SEEN_HEADLINES.add(norm(headline))
    SEEN_URLS.add(url.strip().lower())
    sf = _snippet_fp(snippet)
    if sf and len(sf) > 40:
        SEEN_SNIPPETS.add(sf)


# ─────────────────────────────────────────────────────────────────────────────
# KEYWORD FILTER (Gates 1 & 2)
# ─────────────────────────────────────────────────────────────────────────────

def is_relevant(headline, source, description, category="",
                relevance_kws=None, exclude_kws=None):
    rel_kws = relevance_kws or RELEVANCE_KEYWORDS
    exc_kws = exclude_kws   or EXCLUDE_KEYWORDS
    content = (headline + ' ' + source + ' ' + description).lower()
    for kw in exc_kws:
        if kw in content:
            return False
    if category == "Trade & Logistics":
        if any(kw in content for kw in rel_kws):
            return True
        if any(kw in content for kw in TRADE_IMPACT_KEYWORDS):
            return True
        return False
    return any(kw in content for kw in rel_kws)


# ─────────────────────────────────────────────────────────────────────────────
# RSS  (with retry)
# ─────────────────────────────────────────────────────────────────────────────

def rss_url(query, hours):
    encoded = quote_plus(f"{query} when:{hours}h")
    return f"https://news.google.com/rss/search?q={encoded}&hl=en-IN&gl=IN&ceid=IN:en"


def fetch_feed(query, hours, retry=True):
    """Fetch one RSS feed. Retries once after 3 s on failure."""
    try:
        feed = feedparser.parse(rss_url(query, hours))
        if feed.entries:
            return feed
    except Exception as e:
        if retry:
            time.sleep(3)
            return fetch_feed(query, hours, retry=False)
        print(f"  [ERR] {query}: {e}")
    return None


def article_date(entry):
    try:
        if hasattr(entry, 'published_parsed') and entry.published_parsed:
            return datetime(*entry.published_parsed[:6])
        if hasattr(entry, 'updated_parsed') and entry.updated_parsed:
            return datetime(*entry.updated_parsed[:6])
    except Exception:
        pass
    return datetime.now()


# ─────────────────────────────────────────────────────────────────────────────
# FETCH ALL  — sequential harvest + AI ranking
# ─────────────────────────────────────────────────────────────────────────────

def _harvest_query(query, cat_name, hours, cutoff, per_query_cap,
                   relevance_kws=None, exclude_kws=None):
    """Fetches one RSS feed and returns accepted article dicts."""
    feed = fetch_feed(query, hours)
    if not feed:
        return []

    accepted = []
    for entry in feed.entries[:20]:
        if len(accepted) >= per_query_cap:
            break
        try:
            headline = entry.title        if hasattr(entry, "title")   else ""
            source   = entry.source.title if hasattr(entry, "source")  else "Unknown"
            desc     = entry.summary      if hasattr(entry, "summary") else ""
            date     = article_date(entry)
            link     = entry.link         if hasattr(entry, "link")    else ""
            if not headline or not link:
                continue
            if date < cutoff:
                continue
            if is_irrelevant_source(source, link):
                continue
            if is_duplicate(headline, link, desc):
                continue
            if not is_relevant(headline, source, desc, cat_name,
                               relevance_kws=relevance_kws, exclude_kws=exclude_kws):
                continue
            mark_seen(headline, link, desc)
            accepted.append({
                "headline": headline, "source": source,
                "date":     date.strftime("%Y-%m-%d"),
                "link":     link, "category": cat_name, "desc": desc,
            })
        except Exception:
            continue

    n = len(accepted)
    if n:
        print(f"  [OK]  {query:<70} {n:>3}")
    return accepted


def fetch_all(categories_override=None, relevance_override=None, exclude_override=None):
    hours         = SYSTEM_CONFIG["time_window_hours"]
    cutoff        = datetime.now() - timedelta(hours=hours)
    cat_limit     = SYSTEM_CONFIG["articles_per_category"]
    per_query_cap = SYSTEM_CONFIG.get("articles_per_query", 10)
    min_score     = OLLAMA_CONFIG.get("min_relevance_score", 4)
    results       = defaultdict(list)

    active_cats      = categories_override or CATEGORIES
    active_relevance = relevance_override  or RELEVANCE_KEYWORDS
    active_exclude   = exclude_override    or EXCLUDE_KEYWORDS

    for cat_name, cat_cfg in active_cats.items():
        print(f"\n[{cat_name.upper()}]  ({len(cat_cfg['search_queries'])} queries, sequential)")

        candidates = []
        for q in cat_cfg["search_queries"]:
            try:
                candidates.extend(_harvest_query(q, cat_name, hours, cutoff, per_query_cap,
                                                relevance_kws=active_relevance,
                                                exclude_kws=active_exclude))
            except Exception as e:
                print(f"  [HARVEST ERR] {q}: {e}")

        print(f"  [{cat_name}] {len(candidates)} candidates harvested")
        if not candidates:
            continue

        print(f"  [{cat_name}] Scoring {len(candidates)} candidates with AI…")
        for c in candidates:
            score, reason   = ai_score_article(c["headline"], c["source"], c["desc"], cat_name)
            c["ai_score"]   = score
            c["ai_reason"]  = reason
            if score < min_score:
                print(f"    [DROP] {score}/10 — {c['headline'][:65]}…")

        kept = sorted(
            [c for c in candidates if c["ai_score"] >= min_score],
            key=lambda x: x["ai_score"], reverse=True
        )[:cat_limit]

        print(f"  [{cat_name}] {len(kept)} kept after ranking (top {cat_limit})")

        for c in kept:
            ai_points = ai_summarise_article(
                c["headline"], c["source"], c["desc"], category=cat_name
            )
            print(f"    [OK] {c['ai_score']}/10 — {c['headline'][:60]}…")

            trade_impact_flag = (
                cat_name == "Trade & Logistics" and
                has_trade_impact(c["headline"] + " " + c["desc"])
            )

            results[cat_name].append({
                "headline":      c["headline"],
                "source":        c["source"],
                "date":          c["date"],
                "link":          c["link"],
                "category":      cat_name,
                "ai_score":      c["ai_score"],
                "ai_reason":     c["ai_reason"],
                "ai_points":     ai_points,
                "trade_impact":  trade_impact_flag,
            })

    return results


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(results):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("[EXCEL] openpyxl not installed — skipping Excel export.")
        print("        Run:  pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Intelligence"

    headers = [
        "Category", "Date", "Headline", "Source", "Link",
        "AI Score", "AI Reason", "Trade Impact",
        "Summary 1", "Summary 2", "Summary 3", "Summary 4", "Summary 5"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for cat_name, articles in results.items():
        for a in articles:
            pts = a.get("ai_points", [])
            ws.append([
                cat_name,
                a["date"],
                a["headline"],
                a["source"],
                a["link"],
                a.get("ai_score", ""),
                a.get("ai_reason", ""),
                "Yes" if a.get("trade_impact") else "",
                pts[0] if len(pts) > 0 else "",
                pts[1] if len(pts) > 1 else "",
                pts[2] if len(pts) > 2 else "",
                pts[3] if len(pts) > 3 else "",
                pts[4] if len(pts) > 4 else "",
            ])

    col_widths = [18, 12, 70, 22, 60, 10, 40, 12, 45, 45, 45, 45, 55]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    out = SYSTEM_CONFIG["excel_file"]
    wb.save(out)
    print(f"[EXCEL] Saved → {out}")


# ─────────────────────────────────────────────────────────────────────────────
# FEATURE 4 — SLACK WEBHOOK
# Set SLACK_WEBHOOK_URL in your .env to post a digest to Slack automatically.
# ─────────────────────────────────────────────────────────────────────────────

def post_to_slack(results: dict, industry_label: str = "E-Commerce") -> bool:
    """
    Post a compact digest to Slack via an Incoming Webhook.
    Only the top article per category is included to keep it readable.
    Set SLACK_WEBHOOK_URL in .env to enable.
    """
    webhook_url = os.environ.get("SLACK_WEBHOOK_URL", "").strip()
    if not webhook_url:
        return False
    if not _REQUESTS_OK:
        print("[SLACK] 'requests' package not installed — pip install requests")
        return False

    today  = datetime.now().strftime("%b %d, %Y")
    blocks = [
        {
            "type": "header",
            "text": {"type": "plain_text", "text": f"📡 NewsRadar — {industry_label} Digest · {today}"},
        },
        {"type": "divider"},
    ]

    total = sum(len(v) for v in results.values())
    blocks.append({
        "type": "section",
        "text": {"type": "mrkdwn", "text": f"*{total} articles* curated across {len(results)} categories."},
    })

    for cat_name, articles in results.items():
        if not articles:
            continue
        top = articles[0]   # highest-scored article per category
        score  = top.get("ai_score", "")
        points = top.get("ai_points", [])
        summary_line = points[0] if points else ""
        blocks.append({
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": (
                    f"*{cat_name}* — {len(articles)} article{'s' if len(articles) != 1 else ''}\n"
                    f"Top: <{top['link']}|{top['headline'][:90]}> · AI {score}/10\n"
                    + (f"_{summary_line}_" if summary_line else "")
                ),
            },
        })

    blocks.append({"type": "divider"})
    blocks.append({
        "type": "context",
        "elements": [{"type": "mrkdwn", "text": "Sent by *NewsRadar* — open the full HTML report to review and send the email digest."}],
    })

    try:
        resp = _requests.post(webhook_url, json={"blocks": blocks}, timeout=10)
        if resp.status_code == 200:
            print("[SLACK] Digest posted successfully.")
            return True
        else:
            print(f"[SLACK] Post failed: HTTP {resp.status_code} — {resp.text[:100]}")
            return False
    except Exception as e:
        print(f"[SLACK] Error: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# HTML GENERATION
# ─────────────────────────────────────────────────────────────────────────────

def _summary_html(points: list) -> str:
    if not points:
        return ""
    items = "".join(f"<li>{p}</li>" for p in points)
    return f'<ul class="ai-summary">{items}</ul>'


def article_card(article, color):
    import json as _json
    import base64 as _b64
    h     = article["headline"].replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
    s     = article["source"].replace("&","&amp;")
    link  = article["link"]
    score = article.get("ai_score", "")
    reason_safe  = article.get("ai_reason","").replace('"',"&quot;")
    summary_html = _summary_html(article.get("ai_points",[]))
    card_json  = _json.dumps({
        "headline": article["headline"], "source": article["source"],
        "link": link, "score": score, "reason": article.get("ai_reason",""),
    })
    card_b64 = _b64.b64encode(card_json.encode("utf-8")).decode("ascii")

    return f"""
    <div class="article-card" draggable="true" style="border-left-color:{color}"
         data-date="{article['date']}" data-score="{score}" data-card="{card_b64}">
      <div class="card-left">
        <a href="{link}" target="_blank" class="article-headline">{h}</a>
        <div class="article-meta">
          <span class="meta-source">{s}</span>
          <span class="meta-sep">|</span>
          <span class="meta-date">{article['date']}</span>
          <span class="meta-sep">|</span>
          <span class="meta-score" title="{reason_safe}">AI {score}/10</span>
        </div>
        {summary_html}
        <a href="{link}" target="_blank" class="read-link">Read Full Article</a>
      </div>
      <div class="card-actions">
        <button class="btn-keep"   onclick="markCard(this,'keep')">✓ Keep</button>
        <button class="btn-remove" onclick="deleteCard(this)">✕ Remove</button>
      </div>
    </div>"""


def generate_html(results, categories=None):
    active_cats = categories or CATEGORIES
    today = datetime.now().strftime("%B %d, %Y")
    total = sum(len(arts) for arts in results.values())

    sections_html = ""
    for cat_name, cat_cfg in active_cats.items():
        articles = results.get(cat_name, [])
        count    = len(articles)
        color    = cat_cfg["color"]
        sec_id   = cat_cfg["html_section"]
        desc     = cat_cfg["description"]
        display  = "none" if count > 0 else "block"
        dz_id    = f"dz-{sec_id.replace('s-', '')}"

        cards = "".join(article_card(a, color) for a in articles)

        sections_html += f"""

  <!-- {cat_name.upper()} -->
  <div class="section" id="{sec_id}" data-section="{cat_name}" data-color="{color}">
    <div class="section-header">
      <div class="section-dot" style="background:{color}"></div>
      <div>
        <span class="section-title" style="color:{color}">{cat_name}</span>
        <span class="section-desc">— {desc}</span>
      </div>
      <div class="section-count">{count} article{'s' if count != 1 else ''}</div>
    </div>
    <div class="section-dropzone" id="{dz_id}"></div>
{cards}
    <div class="empty-state" id="empty-{sec_id}" style="display:{display};">No articles in this section.</div>
  </div>
"""

    # FIX 1: The HTML is always served via HTTP (never file://). The browser
    # connects to http://127.0.0.1:5765/ so fetch() to the same origin works
    # in Incognito without any CORS or mixed-content issues.
    #
    # FIX 2: deleteCard() no longer relies solely on transitionend (which can
    # silently fail). It now uses a guaranteed 350ms setTimeout as the primary
    # removal path, with transitionend as an early-exit bonus.
    #
    # FIX 5: Report title is stored in a JS variable (not localStorage, which
    # is blocked in Incognito). sendEmail() reads the live DOM title and sends
    # it to the server in the POST body alongside the serialised HTML.
    #
    # DRAG-DROP v2: Full cross-category drag-and-drop rewrite. Every .section
    # element is now a drop target (not just the narrow dropzone strip at the
    # top). The drag-over logic uses a single document-level handler that finds
    # the nearest card or section boundary, shows a blue insertion line, and
    # inserts the card at the right position on drop. The card's left-border
    # colour and the category label in its meta are both updated to match the
    # destination section.

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>NewsRadar — Industry Intelligence Report</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500;600&display=swap" rel="stylesheet">
<style>
  /* ── FEATURE 3: Dark mode via [data-theme="dark"] on <html> ── */
  :root {{
    --bg:     #f5f6fa;
    --white:  #ffffff;
    --border: #e4e8f0;
    --text:   #111827;
    --muted:  #6b7280;
    --light:  #f3f4f8;
    --ai-bg:  #f8faff;
    --ai-border: #e0e7ff;
  }}
  [data-theme="dark"] {{
    --bg:     #0f1117;
    --white:  #1a1d27;
    --border: #2d3148;
    --text:   #e5e7eb;
    --muted:  #9ca3af;
    --light:  #1f2235;
    --ai-bg:  #1a1f35;
    --ai-border: #2d3a6b;
  }}
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ background:var(--bg); color:var(--text); font-family:'DM Sans',sans-serif; font-size:14px; line-height:1.6; transition:background .2s, color .2s; }}

  .topbar {{ background:var(--white); border-bottom:2px solid var(--border); padding:0 32px; display:flex; align-items:center; justify-content:space-between; height:64px; position:sticky; top:0; z-index:100; box-shadow:0 2px 12px rgba(0,0,0,.05); }}
  .topbar-brand {{ font-family:'Syne',sans-serif; font-size:19px; font-weight:800; color:var(--text); letter-spacing:-.4px; }}
  .topbar-brand span {{ color:#2563eb; }}
  .topbar-sub {{ font-size:11px; color:var(--muted); margin-top:2px; letter-spacing:.5px; text-transform:uppercase; }}
  .topbar-right {{ display:flex; align-items:center; gap:10px; flex-wrap:wrap; }}
  .badge {{ font-family:'DM Mono',monospace; font-size:10px; font-weight:600; padding:5px 12px; border-radius:20px; letter-spacing:.5px; text-transform:uppercase; }}
  .badge-count  {{ background:#eff6ff; color:#2563eb; border:1px solid #bfdbfe; }}

  /* FEATURE 3: dark mode toggle button */
  #dark-toggle {{
    font-family:'DM Mono',monospace; font-size:13px; padding:6px 10px;
    border-radius:20px; border:1px solid var(--border); background:var(--light);
    color:var(--text); cursor:pointer; transition:all .15s; line-height:1;
  }}
  #dark-toggle:hover {{ border-color:#2563eb; color:#2563eb; }}

  .page {{ max-width:1120px; margin:0 auto; padding:32px 24px 80px; }}
  .report-header {{ margin-bottom:16px; }}

  .report-title {{
    font-family:'Syne',sans-serif; font-size:22px; font-weight:800; color:var(--text);
    margin-bottom:4px; cursor:pointer; padding:6px 12px; border-radius:6px;
    transition:background .15s, color .15s; user-select:none; display:inline-block;
  }}
  .report-title:hover {{ background:var(--light); color:#2563eb; }}
  .report-title-hint {{ font-size:10px; color:var(--muted); font-family:'DM Mono',monospace; margin-left:8px; vertical-align:middle; opacity:.6; }}
  #report-title-input {{
    display:none; font-family:'Syne',sans-serif; font-size:22px; font-weight:800;
    border:2px solid #2563eb; border-radius:6px; padding:6px 12px;
    color:var(--text); background:var(--white); width:100%; max-width:700px;
  }}
  .report-meta {{ font-family:'DM Mono',monospace; font-size:12px; color:var(--muted); padding-left:12px; }}

  /* FEATURE 5: search bar */
  .search-wrap {{
    position:relative; margin-bottom:12px;
  }}
  #search-input {{
    width:100%; max-width:480px; padding:9px 16px 9px 38px;
    border:1.5px solid var(--border); border-radius:8px;
    font-family:'DM Sans',sans-serif; font-size:13px;
    background:var(--white); color:var(--text);
    transition:border-color .15s, box-shadow .15s;
    outline:none;
  }}
  #search-input:focus {{ border-color:#2563eb; box-shadow:0 0 0 3px rgba(37,99,235,.12); }}
  .search-icon {{
    position:absolute; left:11px; top:50%; transform:translateY(-50%);
    color:var(--muted); font-size:15px; pointer-events:none;
  }}
  #search-clear {{
    position:absolute; right:12px; top:50%; transform:translateY(-50%);
    background:none; border:none; color:var(--muted); cursor:pointer;
    font-size:16px; display:none; line-height:1;
  }}

  /* FEATURE 2: filter bar */
  .filter-bar {{
    display:flex; gap:8px; flex-wrap:wrap; margin-bottom:20px; align-items:center;
  }}
  .filter-label {{ font-family:'DM Mono',monospace; font-size:11px; color:var(--muted); margin-right:4px; }}
  .filter-group {{ display:flex; gap:4px; }}
  .f-btn {{
    font-family:'DM Mono',monospace; font-size:11px; font-weight:600;
    padding:5px 12px; border-radius:6px; border:1.5px solid var(--border);
    background:var(--white); color:var(--muted); cursor:pointer; transition:all .15s;
  }}
  .f-btn:hover {{ border-color:#2563eb; color:#2563eb; }}
  .f-btn.active {{ background:#2563eb; color:#fff; border-color:#2563eb; }}
  .filter-sep {{ width:1px; background:var(--border); margin:0 4px; }}
  #filter-result-msg {{ font-family:'DM Mono',monospace; font-size:11px; color:var(--muted); margin-left:8px; }}

  .jump-nav {{ display:flex; gap:8px; flex-wrap:wrap; margin-bottom:28px; }}
  .jump-btn {{ background:var(--white); border:1px solid var(--border); color:var(--muted); padding:7px 14px; border-radius:6px; font-size:12px; font-family:'DM Mono',monospace; text-decoration:none; font-weight:500; transition:all .15s; }}
  .jump-btn:hover {{ border-color:#2563eb; color:#2563eb; background:#eff6ff; }}

  .section {{ margin-bottom:36px; }}
  .section-header {{ display:flex; align-items:center; gap:10px; margin-bottom:12px; padding-bottom:10px; border-bottom:2px solid var(--border); }}
  .section-dot {{ width:10px; height:10px; border-radius:50%; flex-shrink:0; }}
  .section-title {{ font-family:'Syne',sans-serif; font-size:15px; font-weight:700; }}
  .section-desc {{ font-size:11px; color:var(--muted); margin-left:4px; }}
  .section-count {{ margin-left:auto; font-family:'DM Mono',monospace; font-size:11px; color:var(--muted); background:var(--light); border:1px solid var(--border); padding:2px 10px; border-radius:10px; }}

  .article-card {{ background:var(--white); border:1px solid var(--border); border-left-width:4px; border-radius:8px; padding:16px 20px; margin-bottom:10px; display:flex; align-items:flex-start; gap:14px; box-shadow:0 1px 3px rgba(0,0,0,.04); transition:box-shadow .15s, opacity .3s, transform .3s; cursor:grab; }}
  .article-card:active {{ cursor:grabbing; }}
  .article-card.dragging {{ opacity:.4; box-shadow:0 8px 32px rgba(0,0,0,.15); transform:scale(1.01); z-index:50; }}
  .article-card:hover {{ box-shadow:0 4px 20px rgba(0,0,0,.08); }}
  .article-card.removing {{ opacity:0; transform:translateX(40px); pointer-events:none; }}
  .article-card.hidden-by-filter {{ display:none !important; }}
  .card-left {{ flex:1; min-width:0; }}
  .article-headline {{ font-size:14px; font-weight:600; color:var(--text); line-height:1.5; margin-bottom:6px; text-decoration:none; display:block; }}
  .article-headline:hover {{ color:#2563eb; text-decoration:underline; }}
  .article-meta {{ display:flex; align-items:center; gap:6px; flex-wrap:wrap; margin-bottom:8px; }}
  .meta-source {{ font-family:'DM Mono',monospace; font-size:11px; color:#2563eb; font-weight:600; }}
  .meta-sep {{ font-size:11px; color:var(--muted); }}
  .meta-date {{ font-size:11px; color:var(--muted); }}
  .meta-score {{ font-family:'DM Mono',monospace; font-size:11px; color:var(--muted); cursor:help; }}

  .ai-summary {{ margin:0 0 10px 0; padding:10px 14px; background:var(--ai-bg); border:1px solid var(--ai-border); border-radius:6px; list-style:none; }}
  .ai-summary li {{ font-size:12.5px; color:var(--text); line-height:1.55; padding:3px 0 3px 14px; position:relative; }}
  .ai-summary li::before {{ content:"▸"; position:absolute; left:0; color:#6366f1; font-size:11px; top:4px; }}

  .read-link {{ display:inline-flex; align-items:center; font-size:12px; font-weight:600; color:#2563eb; text-decoration:none; }}
  .read-link:hover {{ text-decoration:underline; }}
  .read-link::after {{ content:' →'; }}

  .card-actions {{ display:flex; flex-direction:column; gap:6px; flex-shrink:0; }}
  .btn-keep   {{ padding:6px 12px; border-radius:6px; border:1px solid #bbf7d0; background:#f0fdf4; color:#15803d; font-size:11px; font-weight:700; cursor:pointer; font-family:'DM Mono',monospace; transition:all .15s; white-space:nowrap; }}
  .btn-keep:hover   {{ background:#15803d; color:#fff; }}
  .btn-remove {{ padding:6px 12px; border-radius:6px; border:1px solid #fecaca; background:#fef2f2; color:#dc2626; font-size:11px; font-weight:700; cursor:pointer; font-family:'DM Mono',monospace; transition:all .15s; white-space:nowrap; }}
  .btn-remove:hover {{ background:#dc2626; color:#fff; }}

  .section-dropzone {{ min-height:40px; border:2px dashed transparent; border-radius:8px; transition:border-color .2s, background .2s; }}
  .section-dropzone.drop-active {{ border-color:#2563eb; background:#eff6ff; }}
  .section.section-drag-over {{ outline:2px dashed #2563eb; outline-offset:4px; border-radius:10px; background:rgba(37,99,235,.03); }}
  .article-card.drop-above {{ box-shadow:0 -3px 0 0 #2563eb; }}
  .article-card.drop-below {{ box-shadow:0 3px 0 0 #2563eb; }}
  .empty-state {{ text-align:center; padding:32px 20px; color:var(--muted); font-size:13px; background:var(--white); border:1px dashed var(--border); border-radius:8px; }}

  #undo-toast {{ position:fixed; bottom:24px; left:50%; transform:translateX(-50%) translateY(80px); background:#111827; color:#fff; padding:12px 24px; border-radius:8px; font-size:13px; display:flex; align-items:center; gap:14px; box-shadow:0 6px 24px rgba(0,0,0,.3); transition:transform .3s ease; z-index:999; white-space:nowrap; }}
  #undo-toast.visible {{ transform:translateX(-50%) translateY(0); }}
  #undo-btn {{ background:#2563eb; color:#fff; border:none; padding:5px 14px; border-radius:5px; font-size:12px; cursor:pointer; font-family:'DM Mono',monospace; font-weight:600; }}
  #undo-btn:hover {{ background:#1d4ed8; }}
  #drag-label {{ position:fixed; top:-100px; left:-100px; background:#1e3a8a; color:#fff; padding:6px 14px; border-radius:6px; font-size:11px; font-family:'DM Mono',monospace; font-weight:600; pointer-events:none; z-index:9999; white-space:nowrap; }}
  footer {{ border-top:2px solid var(--border); padding:24px 0; text-align:center; font-size:11px; color:var(--muted); font-family:'DM Mono',monospace; background:var(--white); }}
  footer strong {{ color:#2563eb; }}
</style>
</head>
<body>

<div class="topbar">
  <div>
    <div class="topbar-brand">News<span>Radar</span></div>
    <div class="topbar-sub">Weekly Intelligence Digest</div>
  </div>
  <div class="topbar-right">
    <div class="badge badge-count" id="article-count-badge">{total} articles</div>
    <div class="badge" id="feedback-badge" style="background:#f0fdf4;color:#15803d;border:1px solid #bbf7d0;">0 kept · 0 removed</div>
    <button id="dark-toggle" onclick="toggleDark()" title="Toggle dark mode">🌙</button>
    <button onclick="sendEmail()" id="send-email-btn"
      style="font-family:'DM Mono',monospace;font-size:10px;font-weight:700;padding:6px 16px;
             border-radius:20px;border:1px solid #2563eb;background:#2563eb;color:#fff;
             cursor:pointer;letter-spacing:.5px;transition:background .15s;">
      ✉ Send Email
    </button>
  </div>
</div>

<div id="drag-label">Moving article…</div>

<div class="page">

  <div class="report-header">
    <div>
      <span class="report-title" id="report-title-text" onclick="editTitle()"
            title="Click to edit title">Industry Intelligence Report</span>
      <span class="report-title-hint">✎ click to edit</span>
    </div>
    <input type="text" id="report-title-input"
           onkeydown="if(event.key==='Enter') saveTitle(); if(event.key==='Escape') cancelTitle();"
           onblur="saveTitle()">
    <div class="report-meta">{today} &nbsp;·&nbsp; <span id="total-count">{total}</span> articles</div>
  </div>

  <!-- FEATURE 5: Search bar -->
  <div class="search-wrap">
    <span class="search-icon">🔍</span>
    <input type="text" id="search-input" placeholder="Search articles…" oninput="onSearch(this.value)">
    <button id="search-clear" onclick="clearSearch()" title="Clear">✕</button>
  </div>

  <!-- FEATURE 2: Filter bar -->
  <div class="filter-bar">
    <span class="filter-label">SCORE</span>
    <div class="filter-group">
      <button class="f-btn active" data-score="0"  onclick="setScoreFilter(0,this)">All</button>
      <button class="f-btn"        data-score="5"  onclick="setScoreFilter(5,this)">5+</button>
      <button class="f-btn"        data-score="7"  onclick="setScoreFilter(7,this)">7+</button>
      <button class="f-btn"        data-score="9"  onclick="setScoreFilter(9,this)">9+</button>
    </div>
    <div class="filter-sep"></div>
    <span class="filter-label">DATE</span>
    <div class="filter-group">
      <button class="f-btn active" data-date="all"   onclick="setDateFilter('all',this)">All</button>
      <button class="f-btn"        data-date="week"  onclick="setDateFilter('week',this)">This week</button>
      <button class="f-btn"        data-date="today" onclick="setDateFilter('today',this)">Today</button>
    </div>
    <span id="filter-result-msg"></span>
  </div>

  <div class="jump-nav">
    <a href="#s-india"       class="jump-btn">📦 India</a>
    <a href="#s-global"      class="jump-btn">🌐 Global</a>
    <a href="#s-crossborder" class="jump-btn">✈️ Cross-Border &amp; Export</a>
    <a href="#s-trade"       class="jump-btn">🚢 Trade &amp; Logistics</a>
    <a href="#s-competitors" class="jump-btn">⚔️ Competitors</a>
    <a href="#s-others"      class="jump-btn">🔧 Others</a>
  </div>

{sections_html}

</div>

<footer>
  <strong>NewsRadar</strong> — Industry Intelligence Report &nbsp;·&nbsp; v11
</footer>

<div id="undo-toast">
  <span id="undo-msg">Article removed</span>
  <button id="undo-btn" onclick="undoDelete()">UNDO</button>
</div>

<script>
// ════════════════════════════════════════════════════════════════════════════
// STATE
// ════════════════════════════════════════════════════════════════════════════
let _reportTitle = 'Industry Intelligence Report';
const FEEDBACK_PORT = {FEEDBACK_PORT};
const SERVER_URL    = `http://127.0.0.1:${{FEEDBACK_PORT}}`;
let keepCount   = 0;
let removeCount = 0;

// Filter state
let _scoreMin  = 0;
let _dateFilter = 'all';
let _searchTerm = '';

// ════════════════════════════════════════════════════════════════════════════
// FEATURE 3: DARK MODE
// ════════════════════════════════════════════════════════════════════════════
let _darkMode = false;

function toggleDark() {{
  _darkMode = !_darkMode;
  document.documentElement.setAttribute('data-theme', _darkMode ? 'dark' : '');
  document.getElementById('dark-toggle').textContent = _darkMode ? '☀️' : '🌙';
}}

// ════════════════════════════════════════════════════════════════════════════
// FEATURE 5: SEARCH
// ════════════════════════════════════════════════════════════════════════════
function onSearch(val) {{
  _searchTerm = val.trim().toLowerCase();
  document.getElementById('search-clear').style.display = val ? 'block' : 'none';
  applyFilters();
}}

function clearSearch() {{
  document.getElementById('search-input').value = '';
  onSearch('');
}}

// ════════════════════════════════════════════════════════════════════════════
// FEATURE 2: SCORE + DATE FILTERS
// ════════════════════════════════════════════════════════════════════════════
function setScoreFilter(min, btn) {{
  _scoreMin = min;
  btn.closest('.filter-group').querySelectorAll('.f-btn')
     .forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  applyFilters();
}}

function setDateFilter(range, btn) {{
  _dateFilter = range;
  btn.closest('.filter-group').querySelectorAll('.f-btn')
     .forEach(b => b.classList.remove('active'));
  btn.classList.add('active');
  applyFilters();
}}

function applyFilters() {{
  const today = new Date();
  today.setHours(0,0,0,0);
  const weekAgo = new Date(today); weekAgo.setDate(today.getDate() - 7);

  let visible = 0;
  document.querySelectorAll('.article-card').forEach(card => {{
    const score    = parseInt(card.dataset.score || '0', 10);
    const dateStr  = card.dataset.date || '';
    const cardDate = dateStr ? new Date(dateStr) : null;
    const headline = (card.querySelector('.article-headline')?.textContent || '').toLowerCase();
    const source   = (card.querySelector('.meta-source')?.textContent || '').toLowerCase();
    const bullets  = (card.querySelector('.ai-summary')?.textContent || '').toLowerCase();
    const text     = headline + ' ' + source + ' ' + bullets;

    // Score filter
    const scoreOk = score >= _scoreMin;

    // Date filter
    let dateOk = true;
    if (_dateFilter === 'today' && cardDate) {{
      dateOk = cardDate >= today;
    }} else if (_dateFilter === 'week' && cardDate) {{
      dateOk = cardDate >= weekAgo;
    }}

    // Search filter
    const searchOk = !_searchTerm || text.includes(_searchTerm);

    const show = scoreOk && dateOk && searchOk;
    card.classList.toggle('hidden-by-filter', !show);
    if (show) visible++;
  }});

  // Update result message
  const total = document.querySelectorAll('.article-card').length;
  const msg   = document.getElementById('filter-result-msg');
  if (_scoreMin > 0 || _dateFilter !== 'all' || _searchTerm) {{
    msg.textContent = `${{visible}} of ${{total}} shown`;
  }} else {{
    msg.textContent = '';
  }}

  // Update empty states per section
  document.querySelectorAll('.section').forEach(sec => {{
    const visibleInSec = sec.querySelectorAll('.article-card:not(.hidden-by-filter)').length;
    const empty = sec.querySelector('.empty-state');
    const allCards = sec.querySelectorAll('.article-card').length;
    if (empty) empty.style.display = (allCards === 0 || visibleInSec === 0) ? 'block' : 'none';
  }});
}}

// ════════════════════════════════════════════════════════════════════════════
// CROSS-CATEGORY DRAG-AND-DROP  (complete rewrite)
// ════════════════════════════════════════════════════════════════════════════
//
// Design goals:
//   1. Any card can be dragged to any position inside any section.
//   2. A blue insertion line appears ABOVE or BELOW the nearest card,
//      or at the bottom of an empty/sparse section body.
//   3. On drop the card's left-border colour updates to the destination
//      section colour, and the section counts update instantly.
//   4. The drag ghost shows the article headline (truncated).
//   5. Works whether you drop onto a card, between cards, onto the
//      section-dropzone strip, or onto the empty-state placeholder.
//
// Implementation notes:
//   • A single document-level dragover handles ALL positions — no per-card
//     or per-dropzone listeners needed. This avoids the "dead zone" bug
//     where dragging between two cards hits neither card's listener.
//   • _dropTarget stores {{ card, position }} ('before' | 'after') so the
//     drop handler knows exactly where to insert without re-computing.
//   • clearIndicators() strips all visual cues in one pass.
// ════════════════════════════════════════════════════════════════════════════

let _dragSrc     = null;   // the card being dragged
let _dropTarget  = null;   // {{ node, position: 'before'|'after'|'append', section }}

const _dragLabel = document.getElementById('drag-label');

// ── helpers ──────────────────────────────────────────────────────────────────

function clearIndicators() {{
  document.querySelectorAll('.article-card')
    .forEach(c => c.classList.remove('drop-above', 'drop-below', 'dragging'));
  document.querySelectorAll('.section')
    .forEach(s => s.classList.remove('section-drag-over'));
  document.querySelectorAll('.section-dropzone')
    .forEach(dz => dz.classList.remove('drop-active'));
}}

/** Repaint the card's accent colour and update its data-section label. */
function adoptCardToSection(card, section) {{
  const color = section.dataset.color;
  card.style.borderLeftColor = color;
  // Also update the meta-source badge colour to match (subtle visual cue)
  // We do NOT change the card's data-card payload — the original category
  // is kept for feedback logging purposes.
}}

// ── dragstart ────────────────────────────────────────────────────────────────

document.addEventListener('dragstart', e => {{
  const card = e.target.closest('.article-card');
  if (!card) return;
  _dragSrc = card;
  // Defer adding .dragging so Chrome captures the pre-fade ghost image
  requestAnimationFrame(() => card.classList.add('dragging'));
  e.dataTransfer.effectAllowed = 'move';
  const headline = card.querySelector('.article-headline');
  _dragLabel.textContent = (headline ? headline.textContent : 'Article').slice(0, 55) + '…';
  e.dataTransfer.setDragImage(_dragLabel, 0, 0);
}});

// ── dragend ───────────────────────────────────────────────────────────────────

document.addEventListener('dragend', () => {{
  clearIndicators();
  _dragSrc    = null;
  _dropTarget = null;
}});

// ── dragover (single document handler) ───────────────────────────────────────
//
// Priority order for computing _dropTarget:
//   1. If hovering over a non-source article-card → above or below it
//   2. If hovering over a section (but not a card) → append to bottom
//   3. If hovering over a section-dropzone strip → prepend to that section
//
document.addEventListener('dragover', e => {{
  if (!_dragSrc) return;
  e.preventDefault();
  e.dataTransfer.dropEffect = 'move';

  clearIndicators();
  _dropTarget = null;

  // ── Case 1: hovering over another card ───────────────────────────────────
  const overCard = e.target.closest('.article-card');
  if (overCard && overCard !== _dragSrc) {{
    const rect   = overCard.getBoundingClientRect();
    const before = e.clientY < rect.top + rect.height / 2;
    overCard.classList.add(before ? 'drop-above' : 'drop-below');
    _dropTarget = {{
      node:     overCard,
      position: before ? 'before' : 'after',
      section:  overCard.closest('.section'),
    }};
    return;
  }}

  // ── Case 2: hovering over the section-dropzone strip ─────────────────────
  const overDZ = e.target.closest('.section-dropzone');
  if (overDZ) {{
    overDZ.classList.add('drop-active');
    const sec = overDZ.closest('.section');
    _dropTarget = {{ node: overDZ, position: 'prepend', section: sec }};
    return;
  }}

  // ── Case 3: hovering over a section body (empty space / empty-state) ─────
  const overSection = e.target.closest('.section');
  if (overSection) {{
    // Find the closest card to the cursor within this section
    const cards = [...overSection.querySelectorAll('.article-card')]
      .filter(c => c !== _dragSrc);

    if (cards.length === 0) {{
      // Empty section — just mark it as a drop target
      overSection.classList.add('section-drag-over');
      _dropTarget = {{ node: overSection, position: 'append', section: overSection }};
    }} else {{
      // Find which card is nearest vertically
      let nearest = cards[0];
      let minDist = Infinity;
      for (const c of cards) {{
        const rect = c.getBoundingClientRect();
        const mid  = rect.top + rect.height / 2;
        const dist = Math.abs(e.clientY - mid);
        if (dist < minDist) {{ minDist = dist; nearest = c; }}
      }}
      const rect   = nearest.getBoundingClientRect();
      const before = e.clientY < rect.top + rect.height / 2;
      nearest.classList.add(before ? 'drop-above' : 'drop-below');
      _dropTarget = {{
        node:     nearest,
        position: before ? 'before' : 'after',
        section:  overSection,
      }};
    }}
  }}
}});

// ── dragleave — clean up section highlight when cursor leaves ─────────────────

document.addEventListener('dragleave', e => {{
  // Only clear section-level highlight; card indicators are re-drawn on dragover
  const sec = e.target.closest('.section');
  if (sec) sec.classList.remove('section-drag-over');
  const dz = e.target.closest('.section-dropzone');
  if (dz) dz.classList.remove('drop-active');
}});

// ── drop ──────────────────────────────────────────────────────────────────────

document.addEventListener('drop', e => {{
  e.preventDefault();
  if (!_dragSrc || !_dropTarget) {{ clearIndicators(); return; }}

  const {{ node, position, section }} = _dropTarget;
  clearIndicators();

  // Reparent the card
  if (position === 'before') {{
    section.insertBefore(_dragSrc, node);
  }} else if (position === 'after') {{
    node.after(_dragSrc);
  }} else if (position === 'prepend') {{
    // Drop onto the dropzone strip → insert before the first real card,
    // or before the empty-state if the section is empty.
    const firstCard = section.querySelector('.article-card:not(.dragging)');
    const emptyEl   = section.querySelector('.empty-state');
    if (firstCard) {{
      section.insertBefore(_dragSrc, firstCard);
    }} else if (emptyEl) {{
      section.insertBefore(_dragSrc, emptyEl);
    }} else {{
      section.appendChild(_dragSrc);
    }}
  }} else {{
    // 'append' — empty section body drop
    const emptyEl = section.querySelector('.empty-state');
    emptyEl ? section.insertBefore(_dragSrc, emptyEl) : section.appendChild(_dragSrc);
  }}

  // Update card visual to destination section colour
  adoptCardToSection(_dragSrc, section);

  _dragSrc    = null;
  _dropTarget = null;
  refresh();
}});

// ════════════════════════════════════════════════════════════════════════════
// HELPERS
// ════════════════════════════════════════════════════════════════════════════

function cardData(card) {{
  return JSON.parse(atob(card.dataset.card));
}}

async function postFeedback(payload) {{
  try {{
    await fetch(SERVER_URL + '/feedback', {{
      method: 'POST',
      headers: {{'Content-Type': 'application/json'}},
      body: JSON.stringify(payload),
      signal: AbortSignal.timeout(2000),
    }});
  }} catch (_) {{}}
}}

function updateFeedbackBadge() {{
  document.getElementById('feedback-badge').textContent =
    keepCount + ' kept · ' + removeCount + ' removed';
}}

// ════════════════════════════════════════════════════════════════════════════
// KEEP BUTTON
// ════════════════════════════════════════════════════════════════════════════

async function markCard(btn, action) {{
  const card     = btn.closest('.article-card');
  const data     = cardData(card);
  const isToggle = btn.classList.contains('active-' + action);

  card.querySelectorAll('.btn-keep,.btn-remove')
      .forEach(b => b.classList.remove('active-keep','active-remove'));

  if (isToggle) {{
    if (action === 'keep') keepCount   = Math.max(0, keepCount - 1);
    else                   removeCount = Math.max(0, removeCount - 1);
    await postFeedback({{ link: data.link, headline: data.headline,
                          source: data.source, ai_score: data.score, action: null }});
  }} else {{
    btn.classList.add('active-' + action);
    if (action === 'keep') keepCount++;
    else                   removeCount++;
    await postFeedback({{ link: data.link, headline: data.headline,
                          source: data.source, ai_score: data.score, action }});
  }}

  updateFeedbackBadge();
  refresh();
}}

// ════════════════════════════════════════════════════════════════════════════
// REMOVE (✕) BUTTON
// ════════════════════════════════════════════════════════════════════════════

let lastDeleted = null;
let undoTimer   = null;

async function deleteCard(btn) {{
  const card = btn.closest('.article-card');
  const data = cardData(card);

  lastDeleted = {{ card, parent: card.parentNode, next: card.nextSibling, data }};
  removeCount++;
  updateFeedbackBadge();

  await postFeedback({{ link: data.link, headline: data.headline,
                        source: data.source, ai_score: data.score, action: 'remove' }});

  card.classList.add('removing');
  let removed = false;
  const doRemove = () => {{
    if (removed) return;
    removed = true;
    if (card.parentNode) card.remove();
    refresh();
  }};
  card.addEventListener('transitionend', doRemove, {{ once: true }});
  setTimeout(doRemove, 350);

  clearTimeout(undoTimer);
  document.getElementById('undo-msg').textContent = data.headline.slice(0, 60) + '… removed';
  document.getElementById('undo-toast').classList.add('visible');
  undoTimer = setTimeout(() => {{
    document.getElementById('undo-toast').classList.remove('visible');
    lastDeleted = null;
  }}, 5000);
}}

function undoDelete() {{
  if (!lastDeleted) return;
  const {{ card, parent, next }} = lastDeleted;
  card.classList.remove('removing');
  (next && next.parentNode === parent)
    ? parent.insertBefore(card, next)
    : parent.appendChild(card);
  removeCount = Math.max(0, removeCount - 1);
  updateFeedbackBadge();
  refresh();
  document.getElementById('undo-toast').classList.remove('visible');
  lastDeleted = null;
}}

// ════════════════════════════════════════════════════════════════════════════
// COUNT REFRESH
// ════════════════════════════════════════════════════════════════════════════

function refresh() {{
  const total = document.querySelectorAll('.article-card').length;
  document.getElementById('total-count').textContent         = total;
  document.getElementById('article-count-badge').textContent = total + ' articles';
  document.querySelectorAll('.section').forEach(sec => {{
    const count   = sec.querySelectorAll('.article-card').length;
    const counter = sec.querySelector('.section-count');
    if (counter) counter.textContent = count + ' article' + (count !== 1 ? 's' : '');
    const empty = sec.querySelector('.empty-state');
    if (empty) empty.style.display = count === 0 ? 'block' : 'none';
  }});
}}

// ════════════════════════════════════════════════════════════════════════════
// TITLE EDITING
// ════════════════════════════════════════════════════════════════════════════

function editTitle() {{
  const titleEl = document.getElementById('report-title-text');
  const inputEl = document.getElementById('report-title-input');
  inputEl.value = _reportTitle;
  titleEl.style.display = 'none';
  inputEl.style.display = 'block';
  inputEl.focus(); inputEl.select();
}}

function saveTitle() {{
  const titleEl = document.getElementById('report-title-text');
  const inputEl = document.getElementById('report-title-input');
  const val = (inputEl.value || '').trim();
  if (val) {{ _reportTitle = val; titleEl.textContent = val; }}
  titleEl.style.display = 'block';
  inputEl.style.display = 'none';
}}

function cancelTitle() {{
  document.getElementById('report-title-text').style.display = 'block';
  document.getElementById('report-title-input').style.display = 'none';
}}

// ════════════════════════════════════════════════════════════════════════════
// SEND EMAIL
// ════════════════════════════════════════════════════════════════════════════

async function sendEmail() {{
  const btn = document.getElementById('send-email-btn');
  if (!confirm('Send this report by email to all configured recipients?')) return;

  btn.textContent = '⏳ Sending…'; btn.disabled = true;
  btn.style.background = btn.style.borderColor = '#6b7280';

  const resetBtn = () => {{
    btn.textContent = '✉ Send Email'; btn.disabled = false;
    btn.style.background = btn.style.borderColor = '#2563eb';
  }};

  try {{
    const liveHtml = '<!DOCTYPE html>' + document.documentElement.outerHTML;
    const r = await fetch(SERVER_URL + '/send-email', {{
      method: 'POST',
      headers: {{'Content-Type': 'application/json'}},
      body: JSON.stringify({{ title: _reportTitle, html: liveHtml }}),
      signal: AbortSignal.timeout(30000),
    }});
    if (!r.ok) throw new Error(`Server returned HTTP ${{r.status}}`);
    const result = await r.json();
    if (result.ok) {{
      btn.textContent = '✓ Sent!';
      btn.style.background = btn.style.borderColor = '#15803d';
      setTimeout(resetBtn, 4000);
    }} else {{
      throw new Error(result.error || 'SMTP failed — check terminal');
    }}
  }} catch (e) {{
    resetBtn();
    const net = ['TypeError','TimeoutError','AbortError'].includes(e.name) || e.message === 'Failed to fetch';
    alert(net
      ? `⚠️ Could not reach local server on port {FEEDBACK_PORT}.\\n\\nMake sure the terminal running main.py is still open.`
      : '❌ ' + e.message);
  }}
}}
</script>
</body>
</html>"""


# ─────────────────────────────────────────────────────────────────────────────
# EMAIL
# ─────────────────────────────────────────────────────────────────────────────

def _strip_tag_and_contents(html: str, tag_id: str) -> str:
    """
    FIX 3 + FIX 6 — Reliably strip a specific element (by id) and all its
    contents from HTML. Works for arbitrarily nested divs.
    """
    # Find the opening tag that contains the id
    pattern = re.compile(
        r'<([a-z][a-z0-9]*)\b[^>]*\bid=["\']' + re.escape(tag_id) + r'["\'][^>]*>',
        re.IGNORECASE
    )
    m = pattern.search(html)
    if not m:
        return html

    tag_name = m.group(1)
    start    = m.start()
    pos      = m.end()
    depth    = 1
    open_re  = re.compile(r'<' + tag_name + r'\b', re.IGNORECASE)
    close_re = re.compile(r'</' + tag_name + r'\s*>', re.IGNORECASE)

    while depth > 0 and pos < len(html):
        next_open  = open_re.search(html, pos)
        next_close = close_re.search(html, pos)
        if not next_close:
            break
        if next_open and next_open.start() < next_close.start():
            depth += 1
            pos    = next_open.end()
        else:
            depth -= 1
            pos    = next_close.end()

    return html[:start] + html[pos:]


def _make_email_html(review_html: str, report_title: str = "Industry Intelligence Report") -> str:
    """
    FIX 3 — Strip ALL interactive elements properly.
    FIX 5 — Inject the user-edited report title.
    FIX 6 — Email topbar shows ONLY the brand name + date line.
             No article count badge, no feedback badge, no Send Email button.
    """
    h = review_html

    # ── Remove all <script> blocks ────────────────────────────────────────────
    h = re.sub(r'<script\b[^>]*>[\s\S]*?</script>', '', h, flags=re.IGNORECASE)

    # ── Remove interactive UI by element id ──────────────────────────────────
    for eid in ("undo-toast", "drag-label", "send-email-btn",
                "article-count-badge", "feedback-badge"):
        h = _strip_tag_and_contents(h, eid)

    # ── Remove card-actions div (contains Keep / Remove buttons) ─────────────
    # Use a depth-aware stripper for class-based elements
    h = _strip_class_div(h, "card-actions")

    # ── Remove section-dropzone placeholders ──────────────────────────────────
    h = re.sub(r'<div class="section-dropzone"[^>]*>\s*</div>', '', h)

    # ── Remove drag attributes ────────────────────────────────────────────────
    h = re.sub(r'\s+draggable="true"', '', h)

    # ── FIX 5: inject user-edited title ──────────────────────────────────────
    # Replace whatever text is in #report-title-text with the JS-state title
    h = re.sub(
        r'(<span[^>]*id="report-title-text"[^>]*>)[^<]*(</span>)',
        r'\g<1>' + report_title.replace('\\', '').replace('"', '&quot;') + r'\g<2>',
        h,
    )
    # Remove the "✎ click to edit" hint span (class report-title-hint)
    h = re.sub(r'<span class="report-title-hint"[^>]*>.*?</span>', '', h)
    # Remove the hidden title input
    h = re.sub(r'<input[^>]*id="report-title-input"[^>]*>', '', h)

    # ── FIX 6: Strip count/badge/button from topbar-right ────────────────────
    # After the id-based stripping above the topbar-right div may be empty; clean it up
    h = re.sub(r'<div class="topbar-right"[^>]*>\s*</div>', '', h)

    # ── Remove interactive CSS rules that are pointless in email ─────────────
    cursor_rules = [
        r'\.article-card:active\s*\{[^}]*\}',
        r'\.article-card\.dragging\s*\{[^}]*\}',
        r'\.article-card\.drop-above\s*\{[^}]*\}',
        r'\.article-card\.drop-below\s*\{[^}]*\}',
        r'\.article-card\.removing\s*\{[^}]*\}',
        r'\.section-dropzone[^{]*\{[^}]*\}',
        r'#undo-toast[^{]*\{[^}]*\}',
        r'#undo-btn[^{]*\{[^}]*\}',
        r'#drag-label[^{]*\{[^}]*\}',
        r'\.report-title-hint[^{]*\{[^}]*\}',
        r'\.card-actions[^{]*\{[^}]*\}',
        r'\.btn-keep[^{]*\{[^}]*\}',
        r'\.btn-remove[^{]*\{[^}]*\}',
    ]
    for rule in cursor_rules:
        h = re.sub(rule, '', h)

    # ── Make cursor non-grab in email ─────────────────────────────────────────
    h = h.replace('cursor:grab;', '').replace('cursor:grabbing;', '')

    return h


def _strip_class_div(html: str, class_name: str) -> str:
    """
    FIX 3 — Depth-aware stripper for divs matched by class name.
    Handles nested divs correctly where simple regex fails.
    """
    pattern  = re.compile(
        r'<div\b[^>]*\bclass="' + re.escape(class_name) + r'"[^>]*>',
        re.IGNORECASE
    )
    result   = []
    last_pos = 0

    for m in pattern.finditer(html):
        result.append(html[last_pos:m.start()])
        pos   = m.end()
        depth = 1
        open_re  = re.compile(r'<div\b', re.IGNORECASE)
        close_re = re.compile(r'</div\s*>', re.IGNORECASE)

        while depth > 0 and pos < len(html):
            next_open  = open_re.search(html, pos)
            next_close = close_re.search(html, pos)
            if not next_close:
                break
            if next_open and next_open.start() < next_close.start():
                depth += 1
                pos    = next_open.end()
            else:
                depth -= 1
                pos    = next_close.end()

        last_pos = pos

    result.append(html[last_pos:])
    return ''.join(result)


def send_email(html_content: str, label: str = "") -> bool:
    if not EMAIL_CONFIG.get("recipients"):
        print("[EMAIL] No recipients configured.")
        return False

    tag = f" [{label}]" if label else ""
    msg = MIMEMultipart("alternative")
    msg["Subject"] = EMAIL_CONFIG["subject"].format(date=datetime.now().strftime("%Y-%m-%d")) + tag
    msg["From"]    = EMAIL_CONFIG["sender_email"]
    msg["To"]      = ", ".join(EMAIL_CONFIG["recipients"])
    if EMAIL_CONFIG.get("cc_recipients"):
        msg["Cc"]  = ", ".join(EMAIL_CONFIG["cc_recipients"])
    msg.attach(MIMEText(html_content, "html"))

    try:
        with smtplib.SMTP(EMAIL_CONFIG["smtp_server"], EMAIL_CONFIG["smtp_port"]) as s:
            s.starttls()
            s.login(EMAIL_CONFIG["sender_email"], EMAIL_CONFIG["sender_password"])
            all_to = EMAIL_CONFIG["recipients"] + EMAIL_CONFIG.get("cc_recipients", [])
            s.sendmail(EMAIL_CONFIG["sender_email"], all_to, msg.as_string())
        print(f"[EMAIL] Sent to: {', '.join(EMAIL_CONFIG['recipients'])}")
        return True
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")
        return False


def send_from_reviewed_html(html_file: str = "newsradar_report.html") -> None:
    """
    TWO-STAGE EMAIL WORKFLOW
    ========================
    Run:  python main.py --send-email
    Reads the reviewed HTML, strips interactive UI, sends email.
    """
    if not os.path.exists(html_file):
        print(f"[EMAIL] File not found: {html_file}")
        print("        Run the scraper first to generate the report.")
        return

    with open(html_file, "r", encoding="utf-8") as f:
        review_html = f.read()

    print(f"[EMAIL] Read {len(review_html):,} chars from {html_file}")
    email_html = _make_email_html(review_html)
    print(f"[EMAIL] Converted to static email HTML ({len(email_html):,} chars)")
    print(f"[EMAIL] Recipients: {', '.join(EMAIL_CONFIG.get('recipients', []))}")

    _autosave_sent_feedback(email_html)

    confirm = input("[EMAIL] Send now? (yes/no): ").strip().lower()
    if confirm in ("yes", "y"):
        send_email(email_html, label="final")
        print("[LEARN] Feedback updated from sent articles.")
    else:
        print("[EMAIL] Cancelled — no email sent.")


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    if "--send-email" in sys.argv:
        print("=" * 70)
        print("NewsRadar — Send Email from Reviewed Report")
        print("=" * 70)
        send_from_reviewed_html()
        return

    # FEATURE 1: --industry flag
    industry_key = "ecommerce"
    args = sys.argv[1:]
    for i, arg in enumerate(args):
        if arg.startswith("--industry="):
            industry_key = arg.split("=", 1)[1]
        elif arg == "--industry" and i + 1 < len(args):
            industry_key = args[i + 1]

    profile = load_industry_profile(industry_key)
    active_categories = profile["categories"]
    active_relevance  = profile["relevance"]
    active_exclude    = profile["exclude"]

    print("=" * 70)
    print(f"NewsRadar v2.0  —  Industry: {profile['label']}")
    print(f"Time window       : {SYSTEM_CONFIG['time_window_hours']}h")
    print(f"Categories        : {len(active_categories)}")
    print(f"AI enabled        : {OLLAMA_CONFIG.get('enabled', True)}")
    print(f"AI model          : {OLLAMA_CONFIG.get('model', 'qwen3:4b')}")
    print(f"Min AI score      : {OLLAMA_CONFIG.get('min_relevance_score', 7)}/10")
    print(f"Ollama CPU threads: {_cpu_thread_count()} of {os.cpu_count()} logical cores")
    slack_url = os.environ.get("SLACK_WEBHOOK_URL", "")
    print(f"Slack webhook     : {'configured ✓' if slack_url else 'not set (optional)'}")
    fb_exists = os.path.exists(FEEDBACK_FILE)
    print(f"Feedback file     : {FEEDBACK_FILE} ({'found' if fb_exists else 'not found yet'})")
    print("=" * 70)

    _check_ai()
    _load_feedback_examples()

    t0      = time.time()
    results = fetch_all(
        categories_override=active_categories,
        relevance_override=active_relevance,
        exclude_override=active_exclude,
    )
    total   = sum(len(v) for v in results.values())
    elapsed = time.time() - t0
    print(f"\n[DONE] {total} articles across {len(results)} categories in {elapsed:.0f}s")

    html_out = generate_html(results, categories=active_categories)
    out_file = "newsradar_report.html"
    with open(out_file, "w", encoding="utf-8") as f:
        f.write(html_out)
    print(f"[HTML] Saved → {out_file}")

    export_excel(results)

    # FEATURE 4: Slack
    if slack_url:
        post_to_slack(results, industry_label=profile["label"])

    import webbrowser
    server = start_feedback_server()
    url    = f"http://127.0.0.1:{FEEDBACK_PORT}/"

    opened = False
    for browser_name in ("chrome", "google-chrome", "chromium", "chromium-browser"):
        try:
            b = webbrowser.get(browser_name)
            b.open(url)
            opened = True
            print(f"[BROWSER] Opened in Chrome → {url}")
            break
        except Exception:
            pass
    if not opened:
        webbrowser.open(url)
        print(f"[BROWSER] Opened in default browser → {url}")

    print()
    print("─" * 70)
    print("REVIEW WORKFLOW:")
    print("  • Search bar     — filter articles by keyword in real time")
    print("  • Score filter   — show only AI 5+ / 7+ / 9+ articles")
    print("  • Date filter    — today / this week / all")
    print("  • 🌙 Dark mode   — toggle in the topbar")
    print("  • ✓ Keep / ✕ Remove — saved to disk instantly")
    print("  • Drag cards to reorder within or across sections")
    print("  • Click report title to edit before sending")
    print()
    print("WHEN READY TO SEND:")
    print("  Click ✉ Send Email  OR  python main.py --send-email")
    print()
    print("CHANGE INDUSTRY:")
    print("  python main.py --industry fintech")
    print("  python main.py --industry healthcare")
    print("  python main.py --industry tech")
    print("─" * 70)


if __name__ == "__main__":
    main()
